from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models import Transaction
from app.numbers import parse_number, to_fa_digits
from app.schemas import ImportResult


TEMPLATE_HEADERS = [
    "تاریخ",
    "نوع",
    "دسته",
    "عنوان",
    "مبلغ",
    "مرجع",
    "توضیح",
]

SAMPLE_ROWS = [
    ["2026-07-20", "income", "فروش محصول", "فروش لایسنس داشبورد", 28000000, "PRD-003", "نمونه"],
    ["2026-07-20", "expense", "بازاریابی", "تبلیغات اینستاگرام", 3500000, "", "نمونه"],
    ["2026-07-21", "income", "ارائه خدمت", "پشتیبانی ماهانه", 18000000, "SRV-002", "نمونه"],
]


def build_template_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "تراکنش‌ها"
    ws.append(TEMPLATE_HEADERS)
    for row in SAMPLE_ROWS:
        ws.append(row)

    guide = wb.create_sheet("راهنما")
    guide.append(["ستون", "توضیح", "مقادیر مجاز"])
    guide.append(["تاریخ", "تاریخ تراکنش", "YYYY-MM-DD"])
    guide.append(["نوع", "ورودی یا خروجی", "income یا expense یا درآمد یا هزینه"])
    guide.append(["دسته", "دسته‌بندی مالی", "متن آزاد"])
    guide.append(["عنوان", "شرح تراکنش", "متن آزاد"])
    guide.append(["مبلغ", "عدد بدون جداکننده یا با جداکننده", "عدد مثبت"])
    guide.append(["مرجع", "کد محصول/پرسنل اختیاری", "اختیاری"])
    guide.append(["توضیح", "یادداشت", "اختیاری"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _normalize_direction(value: str) -> str | None:
    v = str(value).strip().lower()
    mapping = {
        "income": "income",
        "expense": "expense",
        "درآمد": "income",
        "ورودی": "income",
        "هزینه": "expense",
        "خروجی": "expense",
    }
    return mapping.get(v)


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            return value if hasattr(value, "year") else None
    text = str(value).strip().replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def import_transactions_excel(db: Session, file_path: Path) -> ImportResult:
    wb = load_workbook(filename=str(file_path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    imported = 0
    skipped = 0
    errors: list[str] = []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ImportResult(imported=0, skipped=0, errors=["فایل خالی است"])

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    # Allow Persian or English headers loosely by position
    data_rows = rows[1:]
    for idx, row in enumerate(data_rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            skipped += 1
            continue
        try:
            txn_date = _parse_date(row[0] if len(row) > 0 else None)
            direction = _normalize_direction(row[1] if len(row) > 1 else "")
            category = str(row[2] or "").strip() or "عمومی"
            title = str(row[3] or "").strip() or "تراکنش وارداتی"
            amount_raw = row[4] if len(row) > 4 else 0
            try:
                amount = parse_number(amount_raw if amount_raw not in (None, "") else 0, field="مبلغ")
            except Exception:
                amount = 0.0
            reference = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None
            note = str(row[6]).strip() if len(row) > 6 and row[6] is not None else None

            if not txn_date or not direction or amount <= 0:
                errors.append(f"ردیف {to_fa_digits(idx)}: تاریخ/نوع/مبلغ نامعتبر")
                skipped += 1
                continue

            db.add(
                Transaction(
                    txn_date=txn_date,
                    direction=direction,
                    category=category,
                    title=title,
                    amount=amount,
                    source="excel",
                    reference=reference,
                    note=note,
                )
            )
            imported += 1
        except Exception as exc:
            errors.append(f"ردیف {to_fa_digits(idx)}: {exc}")
            skipped += 1

    db.commit()
    return ImportResult(imported=imported, skipped=skipped, errors=errors[:20])
